import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import viewsets, filters, status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from django.utils.dateparse import parse_date
from django.utils.timezone import now, timedelta
from django.db.models import Sum, F, Value, Max
from django.db.models.functions import Coalesce, Concat
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date
from .models import MedicalCenter, Medicine, Stock, MedicineReceipt, WeeklyConsumptionReport
from .serializers import (
    MedicalCenterSerializer, MedicineSerializer, StockSerializer,
    MedicineReceiptSerializer, WeeklyConsumptionReportSerializer,
    WeeklyReportExcelUploadSerializer,
)
from .filters import StockFilter, WeeklyConsumptionReportFilter, MedicineReceiptFilter

class MedicalCenterViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]

    queryset = MedicalCenter.objects.all()
    serializer_class = MedicalCenterSerializer

class MedicineViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]

    queryset = Medicine.objects.all()
    serializer_class = MedicineSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

class StockViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]
    
    queryset = Stock.objects.select_related('center', 'medicine').all()
    serializer_class = StockSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = StockFilter

class MedicineReceiptViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]

    queryset = MedicineReceipt.objects.select_related('center', 'medicine').all()
    serializer_class = MedicineReceiptSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = MedicineReceiptFilter

class WeeklyConsumptionReportViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]

    queryset = WeeklyConsumptionReport.objects.select_related('center', 'medicine').all()
    serializer_class = WeeklyConsumptionReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = WeeklyConsumptionReportFilter
    search_fields = ['medicine__name']

class WeeklyReportExcelUploadView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser]

    def post(self, request):
        serializer = WeeklyReportExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']

        df = pd.read_excel(file)
        required_columns = ['Date de début de semaine', 'Date de fin de semaine', 'Centre médical', 'Médicament', 'Quantité utilisée']
        if not all(col in df.columns for col in required_columns):
            return Response({"error": "Excel file must contain these columns: " + ", ".join(required_columns)}, status=status.HTTP_400_BAD_REQUEST)

        success_count = 0
        errors = []
        
        # Convert week_start and week_end to datetime if necessary
        df['Date de début de semaine'] = pd.to_datetime(df['Date de début de semaine'], errors='coerce')
        df['Date de fin de semaine'] = pd.to_datetime(df['Date de fin de semaine'], errors='coerce')

        for idx, row in df.iterrows():
            try:
                center_name = row['Centre médical']
                center_name = str(center_name).strip().replace('\u00a0', ' ')

                try:
                    # Try case-insensitive match
                    center = MedicalCenter.objects.get(name__iexact=center_name)
                except MedicalCenter.DoesNotExist:
                    errors.append({"row": idx + 2, "error": f"MedicalCenter '{row['Centre médical']}' does not exist"})
                    continue

                medicine_name = str(row['Médicament']).strip().replace('\u00a0', ' ')
                try:
                    medicine = Medicine.objects.get(name__iexact=medicine_name)
                except Medicine.DoesNotExist:
                    errors.append({"row": idx + 2, "error": f"Medicine '{row['Médicament']}' does not exist"})
                    continue

                # Create WeeklyConsumptionReport
                report = WeeklyConsumptionReport.objects.create(
                    week_start=row['Date de début de semaine'],
                    week_end=row['Date de fin de semaine'],
                    center=center,
                    medicine=medicine,
                    quantity_used=int(row['Quantité utilisée'])
                )
                success_count += 1
            except Exception as e:
                errors.append({"row": idx + 2, "error": str(e)})

        return Response({
            "message": f"Successfully uploaded {success_count} reports.",
            "errors": errors
        }, status=status.HTTP_200_OK)

class MedicineReceiptExcelUploadView(APIView):
    permission_classes = [AllowAny] 
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file_obj)
        except Exception as e:
            return Response({"error": f"Invalid Excel file. Details: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = ['Centre médical', 'Nom du Médicament', 'Unité', 'Quantité Reçue', 'Date de Réception', 'Date de Peramption']
        if not all(col in df.columns for col in required_columns):
            return Response({"error": "Excel file must contain these columns: " + ", ".join(required_columns)}, status=status.HTTP_400_BAD_REQUEST)

        receipts_created = []

        for _, row in df.iterrows():
            try:
                center_name = row['Centre médical']
                center_name = str(center_name).strip().replace('\u00a0', ' ')

                medicine_name = row['Nom du Médicament']
                medicine_name = str(medicine_name).strip().replace('\u00a0', ' ')

                unit = row['Unité']
                quantity = int(row['Quantité Reçue'])
                received_date = pd.to_datetime(row['Date de Réception'], dayfirst=True).date()
                expiration_date = pd.to_datetime(row['Date de Peramption'], dayfirst=True).date()


                try:
                    center = MedicalCenter.objects.get(name=center_name)
                except center.DoesNotExist:
                    center = MedicalCenter.objects.create(name=center_name)

                medicine, _ = Medicine.objects.get_or_create(name=medicine_name, defaults={'unit': unit})

                receipt = MedicineReceipt.objects.create(
                    center=center,
                    medicine=medicine,
                    quantity_received=quantity,
                    exp_date=expiration_date,
                    received_date=received_date
                )

                receipts_created.append(receipt.id)

            except Exception as e:
                return Response({"error": f"Error processing row: {row.to_dict()}. Details: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"message": f"Successfully created {len(receipts_created)} medicine receipts."}, status=status.HTTP_201_CREATED)
    
class WeeklyReportExcelExportView(APIView):
    permission_classes = [AllowAny] 
    def get(self, request, *args, **kwargs):
        # Parse query params
        start_date = parse_date(str(request.query_params.get('start')))
        end_date = parse_date(str(request.query_params.get('end')))
        # Base queryset
        reports = WeeklyConsumptionReport.objects.select_related('medicine', 'center')

        # Apply date filtering if provided
        if start_date and end_date:
            reports = reports.filter(week_end__gte=start_date, week_start__lte=end_date)

        reports = reports.order_by('-week_start')

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Consumption Report"

        # Add title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Weekly Medicine Consumption Report - Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"].style = "Title"

        # Define headers
        headers = [
            "Semaine du", "Semaine au", "Centre médical", "Nom du Médicament",
            "Unité", "Quantité Consommée", "Observation"
        ]
        ws.append(headers)

        # Add data
        for report in reports:
            ws.append([
                report.week_start.strftime('%d-%m-%Y'),
                report.week_end.strftime('%d-%m-%Y'),
                report.center.name,
                report.medicine.name,
                report.medicine.unit,
                report.quantity_used,
                report.observation
            ])

        # Auto-adjust column widths
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Save to memory
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return response
        filename = f"Weekly_Report_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        response = HttpResponse(
            excel_file,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class DashboardAnalyticsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        today = date.today()
        one_month_ago = today - timedelta(days=30)
        four_weeks_ago = today - timedelta(weeks=4)

        # ✅ Summary stats
        summary = {
            "totalMedicines": Medicine.objects.count(),
            "totalCenters": MedicalCenter.objects.count(),
            "totalReceivedQuantity": MedicineReceipt.objects.aggregate(
                total=Coalesce(Sum('quantity_received'), Value(0))
            )['total'],
            "totalStockQuantity": Stock.objects.aggregate(
                total=Coalesce(Sum('total_quantity'), Value(0))
            )['total'],
            "lastReceiptDate": MedicineReceipt.objects.aggregate(
                latest=Coalesce(Max('received_date'), Value(None))
            )['latest']
        }

        # ✅ Weekly consumption per center and medicine (last 4 weeks)
        weekly_consumption = (
            WeeklyConsumptionReport.objects
            .filter(week_start__gte=four_weeks_ago)
            .values(
                center_name=F('center__name'),
                medicine_name=F('medicine__name'),
                unit=F('medicine__unit'),
                weekStart=F('week_start'),
                weekEnd=F('week_end'),
            )
            .annotate(totalUsed=Sum('quantity_used'))
            .order_by('center', 'weekStart')
        )

        # ✅ Group data for frontend: { centerName: [ { period, medicine, totalUsed }... ] }
        grouped_weekly = {}
        for item in weekly_consumption:
            center_name = item.pop('center_name')
            period = f"{item.pop('weekStart')} to {item.pop('weekEnd')}"
            entry = {
                "medicine": item['medicine_name'],
                "unit": item['unit'],
                "totalUsed": item['totalUsed'],
                "period": period,
            }
            grouped_weekly.setdefault(center_name, []).append(entry)

        # ✅ Current stock per center available_stock
        stock_distribution = (
            Stock.objects
            .values(
                center_name=F('center__name'),
                medicine_name=F('medicine__name'),
                available_stock=F('total_quantity'),
                unit=F('medicine__unit'),)
            .annotate(totalQuantity=Sum('total_quantity'))
            .order_by('center')[:5]
        )

        # ✅ Low stock alerts
        low_stock_alerts = (
            Stock.objects
            .filter(total_quantity__lte=10)
            .values(
                center_name=F('center__name'),
                medicine_name=F('medicine__name'),
                unit=F('medicine__unit'),
                stock_quantity=F('total_quantity')
            )
            .order_by('total_quantity')
        )

        # ✅ Recent medicine receipts (latest 5)
        recent_receipts = (
            MedicineReceipt.objects
            .order_by('-received_date')[:5]
            .values(
                center_name=F('center__name'),
                medicine_name=F('medicine__name'),
                unit=F('medicine__unit'),
                quantity__received=F('quantity_received'),
                expiration_date=F('exp_date'),
                received__date=F('received_date'),
            )
        )

        receipts = (
            MedicineReceipt.objects
            .order_by('-received_date')
            .values(
                center_name=F('center__name'),
                medicine_name=F('medicine__name'),
                unit=F('medicine__unit'),
                quantity__received=F('quantity_received'),
                expiration_date=F('exp_date'),
                received__date=F('received_date'),
            )
        )

        # ✅ Top 5 used medicines in last 30 days
        top_used_medicines = (
            WeeklyConsumptionReport.objects
            .filter(week_start__gte=one_month_ago)
            .values(name=F('medicine__name'), unit=F('medicine__unit'))
            .annotate(totalUsed=Sum('quantity_used'))
            .order_by('-totalUsed')[:5]
        )

        return Response({
            "summary": summary,
            "charts": {
                "weeklyConsumptionByCenter": grouped_weekly,
                "topUsedMedicines": list(top_used_medicines),
            },
            "tables": {
                "stockPerCenter": list(stock_distribution),
                "recentReceipts": list(recent_receipts),
                "Receipts": list(receipts),
            },
            "alerts": {
                "lowStock": list(low_stock_alerts),
            }
        })
